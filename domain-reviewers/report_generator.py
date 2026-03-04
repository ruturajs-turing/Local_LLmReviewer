"""
Generate review reports in CSV, Excel, and Markdown formats matching the client's table structure.

Output 1: Per-task CSV  (client Table 1/2 format)
Output 2: Per-task detailed Markdown report
Output 3: Category summary CSV (client Table 3 format)
Output 4: Excel workbook with multiple sheets (Task Scores, QD Details, Category Summary, Stats)
Output 5: Master CSV (auto-appends across runs for easy tracking)
"""

from __future__ import annotations

import csv
import io
import os
import re
from datetime import datetime
from pathlib import Path

from score_schema import CategorySummary, QDResult, TaskScore, aggregate_by_category

QD_IDS = ("A", "B", "C", "E", "F", "G", "H", "J", "M")
QD_NAMES = {
    "A": "Instruction Quality",
    "B": "Tests Quality",
    "C": "Slug & Naming",
    "E": "Dockerfile",
    "F": "test.sh",
    "G": "solve.sh",
    "H": "Task Metadata",
    "J": "Forced Gates",
    "M": "Instruction-Test Alignment",
}


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def _qd_result_map(score: TaskScore) -> dict[str, QDResult | None]:
    """Build a dict mapping QD ID -> QDResult for easy column access."""
    by_id = {}
    for r in score.qd_results:
        by_id[r.qd_id] = r
    return by_id


# ── CSV Reports ────────────────────────────────────────────────────────────

TASK_CSV_HEADERS = [
    "#", "Domain", "Task Name", "Difficulty",
    "QQ (Question Quality)", "TQ (Test Quality)", "TC (Test Coverage)",
    "Overall Avg", "Verdict",
    "QD-A", "QD-B", "QD-C", "QD-E", "QD-F", "QD-G", "QD-H", "QD-J", "QD-M",
    "QD Summary", "Core Issue",
    "QQ Details", "TQ Details", "TC Details",
]


def _task_to_full_row(index: int, s: TaskScore) -> list:
    """Build a comprehensive CSV row with all QD per-column results."""
    qd_map = _qd_result_map(s)
    difficulty = s.details.get("difficulty", "")

    row = [
        index,
        s.category,
        s.task_name,
        difficulty,
        s.qq,
        s.tq,
        s.tc,
        s.overall,
        s.verdict,
    ]
    for qd_id in QD_IDS:
        r = qd_map.get(qd_id)
        if r is None:
            row.append("N/A")
        elif r.passed:
            row.append("PASS")
        else:
            row.append("FAIL")
    row.append(s.qd_summary)
    row.append(s.core_issue)
    row.append(s.details.get("qq_details", "")[:500])
    row.append(s.details.get("tq_details", "")[:500])
    row.append(s.details.get("tc_details", "")[:500])
    return row


def generate_task_csv(scores: list[TaskScore], output_path: str | Path) -> Path:
    """Generate per-task CSV with full QD breakdown per column."""
    path = Path(output_path)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(TASK_CSV_HEADERS)
        for i, score in enumerate(scores, 1):
            writer.writerow(_task_to_full_row(i, score))
    return path


def generate_summary_csv(scores: list[TaskScore], output_path: str | Path) -> Path:
    """Generate category-level summary CSV matching client Table 3 format."""
    summaries = aggregate_by_category(scores)
    path = Path(output_path)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Category", "Count",
            "Avg QQ", "Avg TQ", "Avg TC",
            "Overall Avg", "Core Issue",
        ])
        for s in summaries:
            writer.writerow(s.to_csv_row())
    return path


def append_to_master_csv(scores: list[TaskScore], output_dir: str | Path) -> Path:
    """Append results to a persistent master CSV that accumulates across runs.
    If a task with the same name + domain already exists, it is updated in place.
    """
    path = Path(output_dir) / "master_review_results.csv"
    existing_rows: dict[str, list] = {}
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if len(row) >= 3:
                    key = f"{row[1]}::{row[2]}"
                    existing_rows[key] = row

    for i, s in enumerate(scores, 1):
        key = f"{s.category}::{s.task_name}"
        existing_rows[key] = _task_to_full_row(i, s)

    # Re-index
    sorted_keys = sorted(existing_rows.keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(TASK_CSV_HEADERS)
        for idx, key in enumerate(sorted_keys, 1):
            row = existing_rows[key]
            row[0] = idx
            writer.writerow(row)
    return path


# ── Excel Report ───────────────────────────────────────────────────────────

def generate_excel_report(
    scores: list[TaskScore],
    output_path: str | Path,
    batch_name: str = "",
) -> Path:
    """Generate a comprehensive Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    approved_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    rejected_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    conditional_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")

    def _style_header(ws, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def _auto_width(ws, max_col, max_width=50):
        for col in range(1, max_col + 1):
            width = 10
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        width = max(width, min(len(str(cell.value)), max_width))
            ws.column_dimensions[get_column_letter(col)].width = width + 2

    # ── Sheet 1: Task Scores ──
    ws1 = wb.active
    ws1.title = "Task Scores"
    headers_1 = [
        "#", "Domain", "Task Name", "QQ", "TQ", "TC", "Overall", "Verdict",
        "QD-A", "QD-B", "QD-C", "QD-E", "QD-F", "QD-G", "QD-H", "QD-J", "QD-M",
        "Core Issue",
    ]
    ws1.append(headers_1)
    _style_header(ws1, len(headers_1))

    for i, s in enumerate(scores, 1):
        qd_map = _qd_result_map(s)
        row = [i, s.category, s.task_name, s.qq, s.tq, s.tc, s.overall, s.verdict]
        for qd_id in QD_IDS:
            r = qd_map.get(qd_id)
            row.append("PASS" if r and r.passed else "FAIL" if r else "N/A")
        row.append(s.core_issue[:200])
        ws1.append(row)

        row_num = i + 1
        # Color verdict
        verdict_cell = ws1.cell(row=row_num, column=8)
        if s.verdict == "APPROVED":
            verdict_cell.fill = approved_fill
        elif s.verdict == "REJECTED":
            verdict_cell.fill = rejected_fill
        elif s.verdict == "CONDITIONAL":
            verdict_cell.fill = conditional_fill

        # Color QD cells
        for col_offset, qd_id in enumerate(QD_IDS):
            cell = ws1.cell(row=row_num, column=9 + col_offset)
            r = qd_map.get(qd_id)
            if r and r.passed:
                cell.fill = pass_fill
            elif r and r.is_blocker:
                cell.fill = fail_fill
            else:
                cell.fill = warn_fill

        # Color score cells (QQ, TQ, TC)
        for col in (4, 5, 6):
            cell = ws1.cell(row=row_num, column=col)
            val = cell.value
            if isinstance(val, (int, float)):
                if val >= 7:
                    cell.fill = pass_fill
                elif val >= 5:
                    cell.fill = warn_fill
                else:
                    cell.fill = fail_fill

        for col in range(1, len(headers_1) + 1):
            ws1.cell(row=row_num, column=col).border = thin_border

    _auto_width(ws1, len(headers_1))

    # ── Sheet 2: QD Details ──
    ws2 = wb.create_sheet("QD Details")
    headers_2 = ["#", "Domain", "Task Name", "QD ID", "QD Check", "Result", "Summary", "Issues"]
    ws2.append(headers_2)
    _style_header(ws2, len(headers_2))

    row_idx = 2
    for i, s in enumerate(scores, 1):
        for r in s.qd_results:
            status = "PASS" if r.passed else "FAIL"
            issues_str = " | ".join(r.issues[:5]) if r.issues else ""
            ws2.append([
                i, s.category, s.task_name,
                f"QD-{r.qd_id}", r.qd_name, status,
                (r.summary or "")[:300], issues_str[:500],
            ])
            cell = ws2.cell(row=row_idx, column=6)
            cell.fill = pass_fill if r.passed else fail_fill
            for col in range(1, len(headers_2) + 1):
                ws2.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    _auto_width(ws2, len(headers_2))

    # ── Sheet 3: Category Summary ──
    ws3 = wb.create_sheet("Category Summary")
    summaries = aggregate_by_category(scores)
    headers_3 = ["Category", "Count", "Avg QQ", "Avg TQ", "Avg TC", "Overall Avg", "Core Issue"]
    ws3.append(headers_3)
    _style_header(ws3, len(headers_3))

    for idx, s in enumerate(summaries, 2):
        ws3.append(s.to_csv_row())
        for col in (3, 4, 5, 6):
            cell = ws3.cell(row=idx, column=col)
            val = cell.value
            if isinstance(val, (int, float)):
                if val >= 7:
                    cell.fill = pass_fill
                elif val >= 5:
                    cell.fill = warn_fill
                else:
                    cell.fill = fail_fill
        for col in range(1, len(headers_3) + 1):
            ws3.cell(row=idx, column=col).border = thin_border

    _auto_width(ws3, len(headers_3))

    # ── Sheet 4: Statistics ──
    ws4 = wb.create_sheet("Statistics")
    if scores:
        avg_qq = round(sum(s.qq for s in scores) / len(scores), 1)
        avg_tq = round(sum(s.tq for s in scores) / len(scores), 1)
        avg_tc = round(sum(s.tc for s in scores) / len(scores), 1)
        avg_all = round(sum(s.overall for s in scores) / len(scores), 1)
        approved = sum(1 for s in scores if s.verdict == "APPROVED")
        rejected = sum(1 for s in scores if s.verdict == "REJECTED")
        conditional = sum(1 for s in scores if s.verdict == "CONDITIONAL")

        stats = [
            ("Report", batch_name or "Unnamed"),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Tasks Reviewed", len(scores)),
            ("", ""),
            ("Average QQ", avg_qq),
            ("Average TQ", avg_tq),
            ("Average TC", avg_tc),
            ("Average Overall", avg_all),
            ("", ""),
            ("Approved", f"{approved}/{len(scores)} ({round(approved/len(scores)*100)}%)"),
            ("Conditional", f"{conditional}/{len(scores)} ({round(conditional/len(scores)*100)}%)"),
            ("Rejected", f"{rejected}/{len(scores)} ({round(rejected/len(scores)*100)}%)"),
            ("", ""),
            ("QD Pass Rate", ""),
        ]
        for stat in stats:
            ws4.append(list(stat))

        # QD pass rates
        for qd_id in QD_IDS:
            total = sum(1 for s in scores for r in s.qd_results if r.qd_id == qd_id)
            passed = sum(1 for s in scores for r in s.qd_results if r.qd_id == qd_id and r.passed)
            rate = f"{passed}/{total} ({round(passed/total*100)}%)" if total else "N/A"
            ws4.append([f"QD-{qd_id} ({QD_NAMES.get(qd_id, '')})", rate])

        for col in (1, 2):
            for row in range(1, ws4.max_row + 1):
                ws4.cell(row=row, column=col).border = thin_border
        ws4.column_dimensions["A"].width = 35
        ws4.column_dimensions["B"].width = 30

    # ── Sheet 5: Score Details ──
    ws5 = wb.create_sheet("Score Details")
    headers_5 = ["#", "Domain", "Task Name", "QQ", "QQ Details", "TQ", "TQ Details", "TC", "TC Details", "Core Issue"]
    ws5.append(headers_5)
    _style_header(ws5, len(headers_5))

    for i, s in enumerate(scores, 1):
        ws5.append([
            i, s.category, s.task_name,
            s.qq, s.details.get("qq_details", "")[:500],
            s.tq, s.details.get("tq_details", "")[:500],
            s.tc, s.details.get("tc_details", "")[:500],
            s.core_issue[:300],
        ])
        row_num = i + 1
        for col in range(1, len(headers_5) + 1):
            ws5.cell(row=row_num, column=col).border = thin_border
            ws5.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws5, len(headers_5), max_width=80)

    path = Path(output_path)
    wb.save(path)
    return path


# ── Markdown Report ────────────────────────────────────────────────────────

def _verdict_badge(verdict: str) -> str:
    badges = {
        "APPROVED": "APPROVED",
        "CONDITIONAL": "CONDITIONAL",
        "REJECTED": "REJECTED",
    }
    return badges.get(verdict, verdict)


def generate_markdown_report(
    scores: list[TaskScore],
    output_path: str | Path,
    batch_name: str = "",
) -> Path:
    """Generate detailed Markdown review report."""
    path = Path(output_path)
    lines: list[str] = []

    title = f"Domain-Specific LLM Review Report"
    if batch_name:
        title += f" — {batch_name}"
    lines.append(f"# {title}")
    lines.append(f"\n**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**Tasks Reviewed**: {len(scores)}")
    lines.append("")

    # Summary table (Table 1 format)
    lines.append("## Task Scores\n")
    lines.append("| # | Domain | Task Name | qq | tq | tc | Overall | Verdict | QD Checks | Core Issue |")
    lines.append("|---|--------|-----------|----|----|----|---------|---------|-----------| -----------|")
    for i, s in enumerate(scores, 1):
        lines.append(
            f"| {i} | {s.category} | {s.task_name} | {s.qq} | {s.tq} | {s.tc} "
            f"| {s.overall} | {_verdict_badge(s.verdict)} | {s.qd_summary} | {s.core_issue} |"
        )
    lines.append("")

    # Category summary (Table 3 format)
    summaries = aggregate_by_category(scores)
    if summaries:
        lines.append("## Category Summary\n")
        lines.append("| Category | Count | Question Quality | Test Quality | Test Coverage | Overall Average Score | Core Issue |")
        lines.append("|----------|-------|------------------|--------------|---------------|----------------------|------------|")
        for s in summaries:
            lines.append(
                f"| {s.category} | {s.count} | {s.avg_qq} | {s.avg_tq} | {s.avg_tc} "
                f"| {s.overall_avg} | {s.core_issue} |"
            )
        lines.append("")

    # Statistics
    if scores:
        avg_qq = round(sum(s.qq for s in scores) / len(scores), 1)
        avg_tq = round(sum(s.tq for s in scores) / len(scores), 1)
        avg_tc = round(sum(s.tc for s in scores) / len(scores), 1)
        avg_overall = round(sum(s.overall for s in scores) / len(scores), 1)
        approved = sum(1 for s in scores if s.verdict == "APPROVED")
        rejected = sum(1 for s in scores if s.verdict == "REJECTED")
        conditional = sum(1 for s in scores if s.verdict == "CONDITIONAL")

        lines.append("## Overall Statistics\n")
        lines.append(f"| Metric | Value |")
        lines.append(f"|--------|-------|")
        lines.append(f"| Average qq | {avg_qq} |")
        lines.append(f"| Average tq | {avg_tq} |")
        lines.append(f"| Average tc | {avg_tc} |")
        lines.append(f"| Average Overall | {avg_overall} |")
        lines.append(f"| Approved | {approved}/{len(scores)} |")
        lines.append(f"| Conditional | {conditional}/{len(scores)} |")
        lines.append(f"| Rejected | {rejected}/{len(scores)} |")
        lines.append("")

    # Detailed per-task findings
    lines.append("---\n")
    lines.append("## Detailed Findings\n")

    for i, s in enumerate(scores, 1):
        lines.append(f"### {i}. {s.task_name} ({s.category})\n")
        lines.append(f"**Verdict**: {_verdict_badge(s.verdict)} | **Scores**: qq={s.qq}, tq={s.tq}, tc={s.tc} (overall={s.overall})")
        lines.append(f"\n**Core Issue**: {s.core_issue}\n")

        if s.details:
            if s.details.get("qq_details"):
                lines.append(f"**Question Quality ({s.qq}/10)**:")
                lines.append(f"{s.details['qq_details']}\n")
            if s.details.get("tq_details"):
                lines.append(f"**Test Quality ({s.tq}/10)**:")
                lines.append(f"{s.details['tq_details']}\n")
            if s.details.get("tc_details"):
                lines.append(f"**Test Coverage ({s.tc}/10)**:")
                lines.append(f"{s.details['tc_details']}\n")

        # QD check results
        if s.qd_results:
            lines.append(f"**QD Checks** ({s.qd_summary}):\n")
            lines.append("| QD | Check | Result | Summary |")
            lines.append("|----|-------|--------|---------|")
            for r in s.qd_results:
                badge = "PASS" if r.passed else "FAIL"
                short_summary = (r.summary or "")[:100].replace("|", "/")
                lines.append(f"| {r.qd_id} | {r.qd_name} | {badge} | {short_summary} |")
            lines.append("")

            failed_qds = [r for r in s.qd_results if r.is_blocker]
            if failed_qds:
                lines.append("**Failed QD Details:**\n")
                for r in failed_qds:
                    lines.append(f"- **QD-{r.qd_id} ({r.qd_name})**: {r.summary}")
                    if r.issues:
                        for issue in r.issues[:5]:
                            lines.append(f"  - {issue}")
                lines.append("")

        lines.append("---\n")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def generate_per_task_reports(
    scores: list[TaskScore],
    output_dir: Path,
    batch_name: str = "",
) -> Path:
    """Generate individual per-task review Markdown files that trainers can upload.

    Creates one .md file per task inside outputs/<batch>/per-task/<task-name>.md
    Each file is a self-contained review document for that single task.
    """
    per_task_dir = output_dir / "per-task"
    per_task_dir.mkdir(parents=True, exist_ok=True)

    for s in scores:
        safe_name = re.sub(r'[<>:"/\\|?*]', "_", s.task_name)
        task_path = per_task_dir / f"{safe_name}.md"

        lines: list[str] = []

        # Header
        lines.append(f"# Task Review Report")
        lines.append("")
        lines.append(f"| Field | Value |")
        lines.append(f"|-------|-------|")
        lines.append(f"| **Task** | `{s.task_name}` |")
        lines.append(f"| **Domain** | {s.category} |")
        lines.append(f"| **Difficulty** | {s.details.get('difficulty', 'N/A')} |")
        lines.append(f"| **Reviewed** | {datetime.now().strftime('%Y-%m-%d %H:%M')} |")
        if batch_name:
            lines.append(f"| **Batch** | {batch_name} |")
        lines.append("")

        # Verdict banner
        verdict = s.verdict
        if verdict == "APPROVED":
            verdict_icon = "APPROVED"
        elif verdict == "REJECTED":
            verdict_icon = "REJECTED"
        else:
            verdict_icon = "CONDITIONAL"

        lines.append(f"## Verdict: {verdict_icon}")
        lines.append("")

        # Score card
        lines.append("## Scores")
        lines.append("")
        lines.append("| Dimension | Score | Rating |")
        lines.append("|-----------|:-----:|--------|")
        for label, val in [
            ("Question Quality (qq)", s.qq),
            ("Test Quality (tq)", s.tq),
            ("Test Coverage (tc)", s.tc),
        ]:
            if val >= 7:
                rating = "Good"
            elif val >= 5:
                rating = "Needs Improvement"
            else:
                rating = "Poor"
            lines.append(f"| {label} | **{val}**/10 | {rating} |")
        lines.append(f"| **Overall Average** | **{s.overall}**/10 | |")
        lines.append("")

        # Core issue
        lines.append(f"## Core Issue")
        lines.append("")
        lines.append(f"{s.core_issue}")
        lines.append("")

        # Detailed dimension analysis
        lines.append("## Detailed Analysis")
        lines.append("")

        if s.details.get("qq_details"):
            lines.append(f"### Question Quality ({s.qq}/10)")
            lines.append("")
            lines.append(s.details["qq_details"])
            lines.append("")

        if s.details.get("tq_details"):
            lines.append(f"### Test Quality ({s.tq}/10)")
            lines.append("")
            lines.append(s.details["tq_details"])
            lines.append("")

        if s.details.get("tc_details"):
            lines.append(f"### Test Coverage ({s.tc}/10)")
            lines.append("")
            lines.append(s.details["tc_details"])
            lines.append("")

        # QD checks
        if s.qd_results:
            lines.append("## Quality Dimension (QD) Checks")
            lines.append("")
            lines.append(f"**Summary**: {s.qd_summary}")
            lines.append("")
            lines.append("| QD | Check | Result | Details |")
            lines.append("|----|-------|:------:|---------|")
            for r in s.qd_results:
                badge = "PASS" if r.passed else "FAIL"
                summary = (r.summary or "").replace("|", "/").replace("\n", " ")[:150]
                lines.append(f"| QD-{r.qd_id} | {r.qd_name} | {badge} | {summary} |")
            lines.append("")

            failed_qds = [r for r in s.qd_results if r.is_blocker]
            if failed_qds:
                lines.append("### Failed QD Details")
                lines.append("")
                for r in failed_qds:
                    lines.append(f"**QD-{r.qd_id} — {r.qd_name}**")
                    lines.append("")
                    if r.summary:
                        lines.append(f"{r.summary}")
                        lines.append("")
                    if r.issues:
                        lines.append("Issues found:")
                        lines.append("")
                        for issue in r.issues[:10]:
                            lines.append(f"- {issue}")
                        lines.append("")

        # Recommendations
        lines.append("## Recommendations")
        lines.append("")
        if verdict == "APPROVED":
            lines.append("Task meets quality standards. No changes required.")
        elif verdict == "CONDITIONAL":
            lines.append("Task needs minor improvements before approval:")
            lines.append("")
            if s.qq < 7:
                lines.append(f"- **Improve instruction quality** (current: {s.qq}/10)")
                if s.details.get("qq_details"):
                    first_issue = s.details["qq_details"].split(".")[0]
                    lines.append(f"  - {first_issue}.")
            if s.tq < 7:
                lines.append(f"- **Improve test quality** (current: {s.tq}/10)")
                if s.details.get("tq_details"):
                    first_issue = s.details["tq_details"].split(".")[0]
                    lines.append(f"  - {first_issue}.")
            if s.tc < 7:
                lines.append(f"- **Improve test coverage** (current: {s.tc}/10)")
                if s.details.get("tc_details"):
                    first_issue = s.details["tc_details"].split(".")[0]
                    lines.append(f"  - {first_issue}.")
        else:
            lines.append("Task has critical quality issues and requires significant rework:")
            lines.append("")
            if s.qq <= 5:
                lines.append(f"- **Question quality is critically low** ({s.qq}/10) — instruction needs rewrite")
            if s.tq <= 5:
                lines.append(f"- **Test quality is critically low** ({s.tq}/10) — tests need overhaul")
            if s.tc <= 5:
                lines.append(f"- **Test coverage is critically low** ({s.tc}/10) — major coverage gaps")
            failed_qds = [r for r in s.qd_results if r.is_blocker]
            for r in failed_qds:
                lines.append(f"- **QD-{r.qd_id} ({r.qd_name}) failed** — must be fixed")
        lines.append("")

        # Footer
        lines.append("---")
        lines.append(f"*Generated by Domain-Specific LLM Reviewer (local) — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")

        task_path.write_text("\n".join(lines), encoding="utf-8")

    return per_task_dir


def generate_all_reports(
    scores: list[TaskScore],
    output_dir: str | Path,
    batch_name: str = "",
    fp_results: list[dict] | None = None,
) -> dict[str, Path]:
    """Generate all report formats into organized subfolders.

    Folder structure:
      outputs/<batch>/csv/       — task scores & category summary CSVs
      outputs/<batch>/markdown/  — detailed review report
      outputs/<batch>/excel/     — color-coded Excel workbook
      outputs/<batch>/fp/        — false-positive scan results (if available)
      outputs/master_review_results.csv  — cumulative master CSV
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    ts = _timestamp()
    folder_name = batch_name or "review"

    # Create subfolders
    csv_dir = out / folder_name / "csv"
    md_dir = out / folder_name / "markdown"
    xl_dir = out / folder_name / "excel"
    fp_dir = out / folder_name / "fp"

    csv_dir.mkdir(parents=True, exist_ok=True)
    md_dir.mkdir(parents=True, exist_ok=True)
    xl_dir.mkdir(parents=True, exist_ok=True)
    fp_dir.mkdir(parents=True, exist_ok=True)

    results = {}

    results["task_csv"] = generate_task_csv(
        scores, csv_dir / f"task_scores_{ts}.csv"
    )
    results["summary_csv"] = generate_summary_csv(
        scores, csv_dir / f"category_summary_{ts}.csv"
    )
    results["markdown"] = generate_markdown_report(
        scores, md_dir / f"review_report_{ts}.md", batch_name
    )
    results["excel"] = generate_excel_report(
        scores, xl_dir / f"review_report_{ts}.xlsx", batch_name
    )
    results["master_csv"] = append_to_master_csv(scores, out)

    # Per-task individual reports (one .md per task for trainer upload)
    per_task_base = out / folder_name
    per_task_dir = generate_per_task_reports(scores, per_task_base, batch_name)
    results["per_task_reports"] = per_task_dir

    # FP report (if available)
    if fp_results:
        fp_path = _generate_fp_report(scores, fp_results, fp_dir, ts, batch_name)
        if fp_path:
            results["fp_report"] = fp_path

    return results


def _generate_fp_report(
    scores: list[TaskScore],
    fp_results: list[dict],
    fp_dir: Path,
    ts: str,
    batch_name: str,
) -> Path | None:
    """Generate a false-positive scan report as Markdown + CSV."""
    if not fp_results:
        return None

    # FP Markdown report
    lines = []
    lines.append(f"# False-Positive Scan Report — {batch_name or 'Review'}")
    lines.append(f"\n**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**Tasks Scanned**: {len(fp_results)}")
    lines.append("")

    upgraded = [r for r in fp_results if r.get("fp_verdict") != "KEEP_REJECTED"]
    confirmed = [r for r in fp_results if r.get("fp_verdict") == "KEEP_REJECTED"]
    lines.append(f"**Upgraded**: {len(upgraded)} | **Confirmed**: {len(confirmed)}")
    lines.append("")

    lines.append("## Results\n")
    lines.append("| # | Task | Original Verdict | FP Verdict | QQ | TQ | TC | Adj QQ | Adj TQ | Adj TC | Confidence | Summary |")
    lines.append("|---|------|-----------------|------------|----|----|-----|--------|--------|--------|------------|---------|")

    for i, fp in enumerate(fp_results, 1):
        task_name = fp.get("task_name", "unknown")
        orig_verdict = fp.get("original_verdict", "")
        fp_verdict = fp.get("fp_verdict", "")
        conf = fp.get("confidence", 0)
        lines.append(
            f"| {i} | {task_name} | {orig_verdict} | {fp_verdict} "
            f"| {fp.get('qq_original', '')} | {fp.get('tq_original', '')} | {fp.get('tc_original', '')} "
            f"| {fp.get('qq_adjusted', '')} | {fp.get('tq_adjusted', '')} | {fp.get('tc_adjusted', '')} "
            f"| {conf:.0%} | {fp.get('summary', '')[:100]} |"
        )
    lines.append("")

    # Detailed analysis per task
    if upgraded:
        lines.append("## Upgraded Tasks (False Positives Detected)\n")
        for fp in upgraded:
            lines.append(f"### {fp.get('task_name', 'unknown')}\n")
            lines.append(f"**FP Verdict**: {fp.get('fp_verdict')} (confidence: {fp.get('confidence', 0):.0%})\n")
            if fp.get("mitigating_factors"):
                lines.append("**Mitigating Factors:**")
                for mf in fp["mitigating_factors"]:
                    lines.append(f"- {mf}")
                lines.append("")
            if fp.get("qq_analysis"):
                lines.append(f"**QQ Analysis**: {fp['qq_analysis']}\n")
            if fp.get("tq_analysis"):
                lines.append(f"**TQ Analysis**: {fp['tq_analysis']}\n")
            if fp.get("tc_analysis"):
                lines.append(f"**TC Analysis**: {fp['tc_analysis']}\n")
            lines.append("---\n")

    md_path = fp_dir / f"fp_scan_report_{ts}.md"
    md_path.write_text("\n".join(lines), encoding="utf-8")

    # FP CSV
    csv_path = fp_dir / f"fp_scan_results_{ts}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "#", "Task", "Original Verdict", "FP Verdict",
            "QQ Orig", "TQ Orig", "TC Orig",
            "QQ Adj", "TQ Adj", "TC Adj",
            "Confidence", "Summary",
        ])
        for i, fp in enumerate(fp_results, 1):
            writer.writerow([
                i, fp.get("task_name", ""),
                fp.get("original_verdict", ""), fp.get("fp_verdict", ""),
                fp.get("qq_original", ""), fp.get("tq_original", ""), fp.get("tc_original", ""),
                fp.get("qq_adjusted", ""), fp.get("tq_adjusted", ""), fp.get("tc_adjusted", ""),
                f"{fp.get('confidence', 0):.2f}", fp.get("summary", ""),
            ])

    return md_path
